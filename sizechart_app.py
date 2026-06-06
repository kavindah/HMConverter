#!/usr/bin/env python3
"""
Size Chart Extractor - desktop app
==============================================
Pick a PDF, click Convert, get an Excel size chart with the
standard size run (2XS..4XL) and a grade-difference column after each size.

Run it directly (needs Python + the requirements):
    python sizechart_app.py
or build a standalone Windows .exe / Mac .app with PyInstaller (see README).
"""

import io
import os
import sys
import re
import threading

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===========================================================================
#  EXTRACTION + EXCEL  (same engine as sizechart_to_excel.py)
# ===========================================================================
SIZE_RE = re.compile(r"^(2XS|XS|S|M|L|XL|2XL|3XL|4XL|5XL)(?:-(S|T|TT))?$")
CODE_RE = re.compile(r"^[A-Z][A-Z0-9]*(_[A-Z0-9]+)*$")
BASE_ORDER = ["2XS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL"]
STD_RUN = ["2XS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL"]


def norm(t):
    return re.sub(r"\s+", " ", (t or "")).strip()


def classify_header(header):
    roles, sizes = {}, {}
    for i, cell in enumerate(header):
        c = norm(cell)
        cl = c.lower()
        if not c:
            continue
        if cl.startswith("meas. code") or cl.startswith("meas.code"):
            roles["code"] = i
        elif "point of measurement" in cl:
            roles["name"] = i
        elif "how to measure" in cl:
            roles["howto"] = i
        elif "criticality" in cl:
            roles["crit"] = i
        elif "tolerance" in cl and "-" in cl:
            roles["tolminus"] = i
        elif "tolerance" in cl and "+" in cl:
            roles["tolplus"] = i
        elif SIZE_RE.match(c):
            sizes[i] = c
    return roles, sizes


def parse_banner(cells):
    joined = " ".join(norm(c) for c in cells if c)
    info = {}
    for key, pat in (("set", r"Measurement Set:\s*([\w]+\s*:\s*[\w\-]+)"),
                     ("sample", r"Sample Size:\s*([\w/]+)"),
                     ("uom", r"UOM:\s*([A-Za-z]+)"),
                     ("grade", r"Grade Rule Template:\s*(.+)$")):
        m = re.search(pat, joined)
        if m:
            info[key] = norm(m.group(1))
    return info


def extract(pdf_path, progress=lambda msg: None):
    sets, style = {}, ""

    def get_set(name):
        return sets.setdefault(name, {"order": [], "rec": {}, "meta": {}})

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        for pi, page in enumerate(pdf.pages, 1):
            progress(f"Reading page {pi} of {total}...")
            text = page.extract_text() or ""
            if not style:
                m = re.search(r"Dev Style Name:\s*(.+)", text)
                if m:
                    style = norm(m.group(1))
            for table in page.extract_tables():
                roles, sizes, current = {}, {}, None
                for row in table:
                    if not row:
                        continue
                    first = norm(row[0]).lower()
                    if first.startswith("measurement set:"):
                        info = parse_banner(row)
                        current = info.get("set", current)
                        if current:
                            meta = get_set(current)["meta"]
                            for k in ("sample", "uom", "grade"):
                                if k in info and k not in meta:
                                    meta[k] = info[k]
                        continue
                    if first.startswith("meas. code") or first.startswith("meas.code"):
                        roles, sizes = classify_header(row)
                        continue
                    if "code" not in roles or not sizes:
                        continue
                    code = re.sub(r"\s+", "", row[roles["code"]] or "")
                    if not CODE_RE.match(code):
                        continue
                    if current is None:
                        current = "001 : Measurements"
                    bucket = get_set(current)
                    name = norm(row[roles["name"]]) if "name" in roles else ""
                    namekey = re.sub(r"[^a-z0-9]", "", name.lower())
                    key = (code, namekey)
                    if key not in bucket["rec"]:
                        bucket["order"].append(key)
                        bucket["rec"][key] = {
                            "name": name,
                            "howto": norm(row[roles["howto"]]) if "howto" in roles else "",
                            "crit": norm(row[roles["crit"]]) if "crit" in roles else "",
                            "tolminus": norm(row[roles["tolminus"]]) if "tolminus" in roles else "",
                            "tolplus": norm(row[roles["tolplus"]]) if "tolplus" in roles else "",
                            "vals": {},
                        }
                    rec = bucket["rec"][key]
                    for ci, label in sizes.items():
                        if ci < len(row):
                            v = norm(row[ci])
                            if v:
                                rec["vals"].setdefault(label, v)
    return sets, style


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


FONT = "Arial"
HDR_FILL = PatternFill("solid", fgColor="1F3864")
DIFF_FILL = PatternFill("solid", fgColor="EEF1F7")
HDR_FONT = Font(name=FONT, color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=11)
CELL = Font(name=FONT, size=10)
CODE = Font(name=FONT, size=10, bold=True)
DIFF_FONT = Font(name=FONT, size=9, italic=True, color="7F7F7F")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def write_sheet(wb, set_name, bucket, style, first):
    ws = wb.active if first else wb.create_sheet()
    ws.title = set_name.replace(":", " -")[:31]
    meta = bucket["meta"]
    present = {s for r in bucket["rec"].values() for s in r["vals"]}
    sizes = [s for s in STD_RUN if s in present]
    meta_cols = ["Meas. Code", "Point of Measurement", "How To Measure",
                 "POM Criticality", "Tolerance (-)", "Tolerance (+)"]
    n_meta = len(meta_cols)
    size_col, col = {}, n_meta + 1
    for s in sizes:
        size_col[s] = col
        col += 2
    total_cols = size_col[sizes[-1]] if sizes else n_meta

    banner = (f"Measurement Set: {set_name}    Sample Size: {meta.get('sample','M')}    "
              f"UOM: {meta.get('uom','cm')}    Grade Rule Template: {meta.get('grade','')}").strip()
    ws.cell(row=1, column=1, value=style.strip()).font = TITLE_FONT
    ws.cell(row=2, column=1, value=banner).font = Font(name=FONT, size=9, italic=True)
    span = max(total_cols, n_meta)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=span)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=span)

    hrow = 4
    for c in range(n_meta):
        cell = ws.cell(row=hrow, column=c + 1, value=meta_cols[c])
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    for s in sizes:
        cc = size_col[s]
        cell = ws.cell(row=hrow, column=cc, value=s)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
        if cc + 1 <= total_cols:
            d = ws.cell(row=hrow, column=cc + 1, value="\u0394")
            d.fill = HDR_FILL; d.font = Font(name=FONT, color="FFFFFF", bold=True, size=9)
            d.alignment = CENTER; d.border = BORDER

    r = hrow + 1
    for key in bucket["order"]:
        rec = bucket["rec"][key]
        ws.cell(row=r, column=1, value=key[0]).font = CODE
        ws.cell(row=r, column=2, value=rec["name"]).font = CELL
        ws.cell(row=r, column=3, value=rec["howto"]).font = CELL
        ws.cell(row=r, column=4, value=rec["crit"]).font = CELL
        ws.cell(row=r, column=5, value=num(rec["tolminus"])).font = CELL
        ws.cell(row=r, column=6, value=num(rec["tolplus"])).font = CELL
        for s in sizes:
            cc = size_col[s]
            v = rec["vals"].get(s, "")
            vc = ws.cell(row=r, column=cc, value=num(v) if v != "" else None)
            vc.font = CELL; vc.alignment = CENTER
            if cc + 1 <= total_cols:
                tl, nl = get_column_letter(cc), get_column_letter(cc + 2)
                vc2 = (f'=IF(AND(ISNUMBER({tl}{r}),ISNUMBER({nl}{r})),{nl}{r}-{tl}{r},"")')
                dc = ws.cell(row=r, column=cc + 1, value=vc2)
                dc.font = DIFF_FONT; dc.alignment = CENTER; dc.fill = DIFF_FILL
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c); cell.border = BORDER
            if c in (2, 3):
                cell.alignment = LEFT
        r += 1

    for cl, w in (("A", 11), ("B", 30), ("C", 34), ("D", 13), ("E", 9), ("F", 9)):
        ws.column_dimensions[cl].width = w
    for s in sizes:
        ws.column_dimensions[get_column_letter(size_col[s])].width = 7.5
        if size_col[s] + 1 <= total_cols:
            ws.column_dimensions[get_column_letter(size_col[s] + 1)].width = 5
    ws.freeze_panes = ws.cell(row=hrow + 1, column=n_meta + 1)


def _build(pdf_source, progress=lambda msg: None):
    """Shared builder. pdf_source may be a file path OR a file-like object
    (e.g. BytesIO), so the PDF need never touch disk."""
    sets, style = extract(pdf_source, progress)
    if not sets:
        raise ValueError("No size-chart tables were found in this PDF.")
    progress("Building Excel...")
    wb = Workbook()
    ordered = sorted(sets.keys(), key=lambda s: ("Logo" in s, s))
    for i, name in enumerate(ordered):
        write_sheet(wb, name, sets[name], style, first=(i == 0))
    n = sum(len(sets[x]["order"]) for x in ordered)
    return wb, ordered, n


def convert(pdf_path, out_path, progress=lambda msg: None):
    """Convert a PDF file on disk to an Excel file on disk."""
    wb, ordered, n = _build(pdf_path, progress)
    wb.save(out_path)
    return ordered, n


def convert_to_bytes(pdf_source, progress=lambda msg: None):
    """Convert fully IN MEMORY. pdf_source is a file-like object (BytesIO).
    Returns (xlsx_bytes, sheets, count). Nothing is written to disk."""
    wb, ordered, n = _build(pdf_source, progress)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue(), ordered, n


# ===========================================================================
#  GRAPHICAL APP  (Tkinter - bundled with Python, works on Win/Mac/Linux)
# ===========================================================================
def run_gui():
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    root = tk.Tk()
    root.title(" Size Chart Extractor")
    root.geometry("560x340")
    root.minsize(520, 320)

    pdf_var = tk.StringVar()
    out_var = tk.StringVar()
    status = tk.StringVar(value="Choose a  PDF to begin.")

    pad = {"padx": 14, "pady": 6}
    tk.Label(root, text=" Size Chart Extractor",
             font=("Arial", 15, "bold")).pack(anchor="w", **pad)
    tk.Label(root, text="Pick a PDF, choose where to save the Excel, then Convert.",
             fg="#555").pack(anchor="w", padx=14)

    def pick_pdf():
        p = filedialog.askopenfilename(title="Choose  PDF",
                                       filetypes=[("PDF files", "*.pdf")])
        if p:
            pdf_var.set(p)
            base = os.path.splitext(p)[0]
            out_var.set(base + "_size_chart.xlsx")
            status.set("Ready. Click Convert.")

    def pick_out():
        p = filedialog.asksaveasfilename(title="Save Excel as",
                                         defaultextension=".xlsx",
                                         filetypes=[("Excel files", "*.xlsx")])
        if p:
            out_var.set(p)

    f1 = tk.Frame(root); f1.pack(fill="x", **pad)
    tk.Entry(f1, textvariable=pdf_var).pack(side="left", fill="x", expand=True)
    tk.Button(f1, text="Choose PDF...", command=pick_pdf).pack(side="left", padx=6)

    f2 = tk.Frame(root); f2.pack(fill="x", padx=14)
    tk.Entry(f2, textvariable=out_var).pack(side="left", fill="x", expand=True)
    tk.Button(f2, text="Save as...", command=pick_out).pack(side="left", padx=6)

    bar = ttk.Progressbar(root, mode="indeterminate")
    bar.pack(fill="x", padx=14, pady=10)
    tk.Label(root, textvariable=status, fg="#333", wraplength=520,
             justify="left").pack(anchor="w", padx=14)

    def set_status(msg):
        status.set(msg)
        root.update_idletasks()

    def do_convert():
        pdf, out = pdf_var.get(), out_var.get()
        if not pdf:
            messagebox.showwarning("No PDF", "Please choose a PDF first.")
            return
        if not out:
            out = os.path.splitext(pdf)[0] + "_size_chart.xlsx"
            out_var.set(out)
        convert_btn.config(state="disabled")
        bar.start(12)

        def worker():
            try:
                sheets, n = convert(pdf, out, progress=set_status)
                set_status(f"Done. {n} measurements across {len(sheets)} sheet(s).")
                messagebox.showinfo("Finished", f"Saved:\n{out}")
                try:
                    folder = os.path.dirname(out) or "."
                    if sys.platform.startswith("win"):
                        os.startfile(folder)            # noqa
                    elif sys.platform == "darwin":
                        os.system(f'open "{folder}"')
                except Exception:
                    pass
            except Exception as e:
                set_status("Error: " + str(e))
                messagebox.showerror("Could not convert", str(e))
            finally:
                bar.stop()
                convert_btn.config(state="normal")

        threading.Thread(target=worker, daemon=True).start()

    convert_btn = tk.Button(root, text="Convert  →  Excel", height=2,
                            bg="#1F3864", fg="white", font=("Arial", 11, "bold"),
                            command=do_convert)
    convert_btn.pack(fill="x", padx=14, pady=12)

    root.mainloop()


def main():
    # Command-line mode:  python sizechart_app.py input.pdf [output.xlsx]
    if len(sys.argv) >= 2 and sys.argv[1].lower().endswith(".pdf"):
        pdf = sys.argv[1]
        out = sys.argv[2] if len(sys.argv) >= 3 else os.path.splitext(pdf)[0] + "_size_chart.xlsx"
        sheets, n = convert(pdf, out, progress=lambda m: print(m))
        print(f"Saved -> {out}  ({n} measurements, {len(sheets)} sheet(s))")
    else:
        run_gui()


if __name__ == "__main__":
    main()
